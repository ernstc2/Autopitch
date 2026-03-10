from __future__ import annotations

import re
from pathlib import Path
from typing import BinaryIO, Union

import openpyxl

from autopitch.models import FinancialData, StatementData, ValidationError

REQUIRED_SHEETS = ["P&L", "Balance Sheet", "Cash Flow"]

SHEET_TO_FIELD = {
    "P&L": "pl",
    "Balance Sheet": "balance_sheet",
    "Cash Flow": "cash_flow",
}


def find_year_columns(ws) -> dict[str, int]:
    """Scan row 1 for FY#### headers and return {year_label: column_index}."""
    year_cols: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value and re.match(r"FY\d{4}", str(cell.value)):
            year_cols[str(cell.value)] = cell.column
    return year_cols


def extract_rows(ws, year_cols: dict[str, int]) -> dict[str, dict[str, float | None]]:
    """Extract row data from worksheet using dynamic year columns."""
    rows: dict[str, dict[str, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if label is None:
            continue  # skip blank/merged continuation rows
        values: dict[str, float | None] = {}
        for year, col_idx in year_cols.items():
            raw = row[col_idx - 1]
            values[year] = float(raw) if raw is not None else None
        rows[str(label).strip()] = values
    return rows


def parse_workbook(source: Union[str, Path, BinaryIO]) -> FinancialData:
    """Parse an Excel workbook into a FinancialData object.

    Args:
        source: File path (str or Path) or binary stream (BinaryIO).

    Returns:
        FinancialData populated from the workbook's three required sheets.

    Raises:
        ValidationError: If required sheets are missing or validation fails.
    """
    # Determine company name before opening (path gives stem; stream gives "Unknown")
    if isinstance(source, (str, Path)):
        company_name = Path(source).stem
    else:
        company_name = "Unknown"

    wb = openpyxl.load_workbook(source, data_only=True)

    # Validate required sheets are present — collect all missing before raising
    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        errors = [
            f"Missing required sheet: '{sheet}'. Workbook must contain sheets named "
            f"'P&L', 'Balance Sheet', and 'Cash Flow'."
            for sheet in missing_sheets
        ]
        raise ValidationError(errors)

    # Build statement data for each sheet
    statements: dict[str, StatementData] = {}
    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        year_cols = find_year_columns(ws)
        rows = extract_rows(ws, year_cols)
        years = list(year_cols.keys())
        statements[SHEET_TO_FIELD[sheet_name]] = StatementData(years=years, rows=rows)

    data = FinancialData(
        company_name=company_name,
        pl=statements["pl"],
        balance_sheet=statements["balance_sheet"],
        cash_flow=statements["cash_flow"],
    )

    # Validation gate — must pass before returning
    from autopitch.validator import validate  # noqa: PLC0415 (deferred to avoid circular)
    validate(data)

    return data
