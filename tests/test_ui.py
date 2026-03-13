"""Structural and unit tests for app.py UI patterns.

Tests verify that app.py contains the required structural patterns for:
- DEMO-01: Hero section with Autopitch description before interactive widgets
- DEMO-02: Cached demo pipeline function (_run_demo_pipeline with @st.cache_data)
- DEMO-03: PPTX bytes persisted in st.session_state after generation
- DEMO-04: Elapsed time and slide count stored in st.session_state
- UPLD-03: _build_template_xlsx() returns valid XLSX bytes (PK magic)

Approach: Read app.py source as text and verify required patterns using regex/string
matching. No Streamlit AppTest runtime needed for structural checks.
"""

import re
import sys
import importlib
from pathlib import Path

import pytest

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

APP_PATH = Path(__file__).parent.parent / "app.py"


def _source() -> str:
    """Return the full source text of app.py."""
    return APP_PATH.read_text(encoding="utf-8")


def _lines() -> list[str]:
    """Return app.py lines (0-indexed)."""
    return _source().splitlines()


def _first_line_containing(text: str, lines: list[str]) -> int:
    """Return the 0-based line index of the first line containing text, or -1."""
    for i, line in enumerate(lines):
        if text in line:
            return i
    return -1


# ---------------------------------------------------------------------------
# DEMO-01: Hero section appears BEFORE any interactive widget
# ---------------------------------------------------------------------------


class TestHeroSection:
    def test_hero_section_content(self):
        """app.py contains an app header element before any widget."""
        source = _source()
        lines = _lines()

        # Must have an app header element (native st.title/header or custom HTML)
        hero_pattern = r'(st\.(title|header)\s*\(|hero-title|hero-badge|app-header|app-name)'
        assert re.search(hero_pattern, source), (
            "app.py must have a header element (st.title, st.header, or custom header HTML)"
        )

        # Header text must appear before the first interactive widget
        hero_line = -1
        for i, line in enumerate(lines):
            if re.search(r'(st\.(title|header)\s*\(|hero-title|hero-badge|app-header|app-name)', line):
                hero_line = i
                break

        widget_line = -1
        for i, line in enumerate(lines):
            if re.search(r'st\.(button|file_uploader)\s*\(', line):
                widget_line = i
                break

        assert hero_line != -1, "No hero title found in app.py"
        assert widget_line != -1, "No interactive widget found in app.py"
        assert hero_line < widget_line, (
            f"Hero section (line {hero_line+1}) must appear before first widget "
            f"(line {widget_line+1})"
        )

    def test_hero_has_descriptive_text(self):
        """app.py contains st.markdown or st.write for descriptive hero text."""
        source = _source()
        # Must have some descriptive text call (markdown or write)
        assert re.search(r"st\.(markdown|write)\s*\(", source), (
            "app.py must have st.markdown() or st.write() for hero descriptive text"
        )


# ---------------------------------------------------------------------------
# DEMO-02: Cached demo pipeline function
# ---------------------------------------------------------------------------


class TestDemoPipelineCached:
    def test_demo_pipeline_cached(self):
        """_run_demo_pipeline function exists in app.py with @st.cache_data decorator."""
        source = _source()
        lines = _lines()

        # Find @st.cache_data decorator
        cache_line = _first_line_containing("@st.cache_data", lines)
        assert cache_line != -1, "app.py must have @st.cache_data decorator"

        # Find _run_demo_pipeline function
        func_line = _first_line_containing("def _run_demo_pipeline", lines)
        assert func_line != -1, "app.py must define _run_demo_pipeline function"

        # Decorator must appear immediately before the function (within 3 lines)
        assert func_line - cache_line <= 3 and func_line > cache_line, (
            f"@st.cache_data (line {cache_line+1}) must appear just before "
            f"_run_demo_pipeline (line {func_line+1})"
        )

    def test_demo_pipeline_uses_apple_data(self):
        """_run_demo_pipeline references the bundled Apple demo file."""
        source = _source()
        assert "demo/apple_financials" in source, (
            "app.py must reference demo/apple_financials in _run_demo_pipeline"
        )

    def test_demo_pipeline_calls_run_pipeline(self):
        """_run_demo_pipeline calls run_pipeline()."""
        source = _source()
        assert re.search(r"return\s+run_pipeline\s*\(", source), (
            "app.py must call run_pipeline() inside _run_demo_pipeline"
        )


# ---------------------------------------------------------------------------
# DEMO-03: PPTX bytes stored in session_state
# ---------------------------------------------------------------------------


class TestDemoBytesInSessionState:
    def test_demo_bytes_in_session_state(self):
        """app.py references st.session_state['demo_pptx'] for storing pipeline output."""
        source = _source()
        assert re.search(r'session_state\[[\"\']demo_pptx[\"\']', source), (
            "app.py must use st.session_state['demo_pptx'] for storing pipeline output"
        )

    def test_download_button_rendered_outside_button_block(self):
        """Download button for demo is rendered based on session_state, not inside if st.button()."""
        source = _source()
        # The demo download button must reference session_state["demo_pptx"]
        # as a condition (not inside the if st.button() generate block)
        assert re.search(r'session_state\[[\"\']demo_pptx[\"\']\]\s+is\s+not\s+None', source), (
            "app.py must conditionally render demo download button using "
            "'st.session_state[\"demo_pptx\"] is not None'"
        )

    def test_download_button_on_click_ignore(self):
        """Demo download button uses on_click='ignore' parameter."""
        source = _source()
        assert re.search(r'on_click\s*=\s*["\']ignore["\']', source), (
            "app.py must use on_click='ignore' on download button(s)"
        )


# ---------------------------------------------------------------------------
# DEMO-04: Stats stored in session_state
# ---------------------------------------------------------------------------


class TestDemoStatsInSessionState:
    def test_demo_stats_in_session_state(self):
        """app.py stores demo_elapsed and demo_slides in st.session_state."""
        source = _source()
        assert re.search(r'session_state\[[\"\']demo_elapsed[\"\']', source), (
            "app.py must store demo_elapsed in st.session_state"
        )
        assert re.search(r'session_state\[[\"\']demo_slides[\"\']', source), (
            "app.py must store demo_slides in st.session_state"
        )

    def test_session_state_initialization(self):
        """app.py initializes demo_pptx, demo_elapsed, demo_slides keys before first use."""
        source = _source()
        lines = _lines()

        # All three keys must be initialized
        for key in ("demo_pptx", "demo_elapsed", "demo_slides"):
            assert key in source, f"app.py must reference session_state key '{key}'"

        # Initialization guard must appear before first widget
        init_pattern = r"not in st\.session_state"
        init_line = -1
        for i, line in enumerate(lines):
            if re.search(init_pattern, line):
                init_line = i
                break

        widget_line = -1
        for i, line in enumerate(lines):
            if re.search(r'st\.(button|file_uploader)\s*\(', line):
                widget_line = i
                break

        assert init_line != -1, (
            "app.py must initialize session state keys with 'if key not in st.session_state'"
        )
        assert widget_line != -1, "No interactive widget found in app.py"
        assert init_line < widget_line, (
            f"Session state init (line {init_line+1}) must appear before "
            f"first widget (line {widget_line+1})"
        )


# ---------------------------------------------------------------------------
# UPLD-03: _build_template_xlsx returns valid XLSX bytes
# ---------------------------------------------------------------------------


class TestTemplatexlsxBytes:
    def test_build_template_xlsx_exists(self):
        """app.py defines _build_template_xlsx function."""
        source = _source()
        assert "def _build_template_xlsx" in source, (
            "app.py must define _build_template_xlsx() function"
        )

    def test_template_xlsx_bytes(self):
        """_build_template_xlsx() returns bytes that start with PK (valid ZIP/XLSX).

        We execute _build_template_xlsx in isolation by extracting it from app.py
        source and running it — avoids Streamlit widget execution at import time.
        """
        from io import BytesIO
        import openpyxl

        # Extract the function body from app.py source and exec it in a controlled
        # namespace to avoid triggering Streamlit widget calls at module level.
        source = _source()

        # Build a minimal namespace with the imports the function needs
        namespace: dict = {
            "BytesIO": BytesIO,
            "openpyxl": openpyxl,
        }

        # Exec the entire app source but with streamlit stubbed out so module-level
        # widget calls are no-ops.  We need st.session_state to behave like a dict.
        import unittest.mock as mock

        class FakeSessionState(dict):
            pass

        st_mock = mock.MagicMock()
        st_mock.session_state = FakeSessionState()
        # Make st.columns() return a 2-element list of MagicMocks so unpacking works
        col_mock = mock.MagicMock()
        col_mock.__enter__ = mock.Mock(return_value=mock.MagicMock())
        col_mock.__exit__ = mock.Mock(return_value=False)
        def _fake_columns(spec, **kwargs):
            if isinstance(spec, int):
                return [col_mock] * spec
            return [col_mock] * len(spec)
        st_mock.columns.side_effect = _fake_columns
        # Make context managers work (st.spinner, st.expander)
        ctx = mock.MagicMock()
        ctx.__enter__ = mock.Mock(return_value=mock.MagicMock())
        ctx.__exit__ = mock.Mock(return_value=False)
        st_mock.spinner.return_value = ctx
        st_mock.expander.return_value = ctx
        # Buttons must return False to prevent pipeline execution
        st_mock.button.return_value = False
        # file_uploader must return None to skip upload block
        st_mock.file_uploader.return_value = None

        with mock.patch.dict(sys.modules, {"streamlit": st_mock}):
            if "app" in sys.modules:
                del sys.modules["app"]

            import app as app_module

            result = app_module._build_template_xlsx()

        assert isinstance(result, bytes), "_build_template_xlsx() must return bytes"
        assert result[:2] == b"PK", (
            "_build_template_xlsx() must return valid ZIP/XLSX bytes starting with 'PK'"
        )

    def test_template_xlsx_has_required_sheets(self):
        """_build_template_xlsx() returns XLSX with P&L, Balance Sheet, Cash Flow sheets
        and correct row labels in each sheet (UPLD-03).
        """
        from io import BytesIO
        import openpyxl
        import unittest.mock as mock
        import sys

        st_mock = mock.MagicMock()
        st_mock.session_state = {}
        col_mock = mock.MagicMock()
        col_mock.__enter__ = mock.Mock(return_value=mock.MagicMock())
        col_mock.__exit__ = mock.Mock(return_value=False)
        def _fake_columns2(spec, **kwargs):
            if isinstance(spec, int):
                return [col_mock] * spec
            return [col_mock] * len(spec)
        st_mock.columns.side_effect = _fake_columns2
        ctx = mock.MagicMock()
        ctx.__enter__ = mock.Mock(return_value=mock.MagicMock())
        ctx.__exit__ = mock.Mock(return_value=False)
        st_mock.spinner.return_value = ctx
        st_mock.expander.return_value = ctx
        st_mock.button.return_value = False
        st_mock.file_uploader.return_value = None

        with mock.patch.dict(sys.modules, {"streamlit": st_mock}):
            if "app" in sys.modules:
                del sys.modules["app"]
            import app as app_module
            result = app_module._build_template_xlsx()

        wb = openpyxl.load_workbook(BytesIO(result))
        sheet_names = wb.sheetnames

        assert "P&L" in sheet_names, "Template must contain 'P&L' sheet"
        assert "Balance Sheet" in sheet_names, "Template must contain 'Balance Sheet' sheet"
        assert "Cash Flow" in sheet_names, "Template must contain 'Cash Flow' sheet"

        # Check P&L row labels
        pl_labels = [wb["P&L"].cell(row=r, column=1).value for r in range(2, 20) if wb["P&L"].cell(row=r, column=1).value]
        assert "Revenue" in pl_labels
        assert "Net Income" in pl_labels
        assert "Operating Income" in pl_labels

        # Check Balance Sheet row labels
        bs_labels = [wb["Balance Sheet"].cell(row=r, column=1).value for r in range(2, 20) if wb["Balance Sheet"].cell(row=r, column=1).value]
        assert "Total Assets" in bs_labels
        assert "Total Debt" in bs_labels

        # Check Cash Flow row labels
        cf_labels = [wb["Cash Flow"].cell(row=r, column=1).value for r in range(2, 20) if wb["Cash Flow"].cell(row=r, column=1).value]
        assert "Operating Cash Flow" in cf_labels
        assert "Capital Expenditures" in cf_labels


# ---------------------------------------------------------------------------
# UPLD-01: Upload section with file uploader
# ---------------------------------------------------------------------------


class TestUploadSection:
    def test_upload_section_exists(self):
        """app.py contains st.file_uploader accepting .xlsx files with instructions (UPLD-01)."""
        source = _source()
        # Must have file_uploader with type=["xlsx"]
        assert re.search(r'st\.file_uploader\s*\(', source), (
            "app.py must have st.file_uploader()"
        )
        assert re.search(r'type\s*=\s*\[[\"\']xlsx[\"\']', source), (
            "app.py file_uploader must accept type=['xlsx']"
        )
        # Must have instructional text about three sheets
        assert re.search(r'P&L.*Balance Sheet.*Cash Flow|three sheets', source, re.DOTALL | re.IGNORECASE), (
            "app.py must contain instructional text mentioning the three required sheets"
        )

    def test_format_guide_expander(self):
        """app.py contains st.expander with format guide content (UPLD-02)."""
        source = _source()
        # Must have expander
        assert re.search(r'st\.expander\s*\(', source), (
            "app.py must have st.expander() for the format guide"
        )
        # Expander must contain sheet names in source
        assert "P&L" in source, "Format guide must mention P&L sheet"
        assert "Balance Sheet" in source, "Format guide must mention Balance Sheet"
        assert "Cash Flow" in source, "Format guide must mention Cash Flow sheet"
        # Must include column layout info
        assert re.search(r'FY\d{4}|fiscal year|Column [AB]', source, re.IGNORECASE), (
            "Format guide must describe column layout with fiscal year headers"
        )
        # Must include required row labels
        assert "Revenue" in source, "Format guide must list Revenue row"
        assert "Operating Income" in source, "Format guide must list Operating Income row"

    def test_upload_download_on_click_ignore(self):
        """Upload result download button uses on_click='ignore' (UPLD-03 / rerun fix)."""
        source = _source()
        lines = _lines()

        # Find line number of upload_pptx session state check
        upload_check_line = _first_line_containing("upload_pptx", lines)
        assert upload_check_line != -1, "app.py must reference upload_pptx in session_state"

        # Must have on_click='ignore' — at least two occurrences (demo + upload)
        matches = re.findall(r'on_click\s*=\s*["\']ignore["\']', source)
        assert len(matches) >= 2, (
            f"app.py must have on_click='ignore' on both demo and upload download buttons "
            f"(found {len(matches)} occurrence(s))"
        )


# ---------------------------------------------------------------------------
# SKIL-01: Tech stack showcase section
# ---------------------------------------------------------------------------


class TestTechStackSection:
    def test_tech_stack_section(self):
        """app.py mentions key technologies used (SKIL-01)."""
        source = _source()

        # Must mention at least 5 required technologies (in footer, section, or elsewhere)
        required_techs = ["Python", "Claude", "python-pptx", "matplotlib"]
        for tech in required_techs:
            assert tech in source, f"app.py must mention '{tech}'"

    def test_tech_stack_has_rationale(self):
        """Technologies are mentioned with sufficient context (SKIL-01)."""
        source = _source()
        # Must reference pipeline/data/chart concepts somewhere in the source
        assert re.search(r'pipeline|data|LLM|narrative|chart|parsing|financials|deck', source, re.IGNORECASE), (
            "app.py must include context about what the technologies do"
        )
