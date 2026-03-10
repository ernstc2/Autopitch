import openpyxl
from io import BytesIO
from pathlib import Path
import pytest
from autopitch.parser import parse_workbook
from autopitch.models import FinancialData, ValidationError


def _build_minimal_workbook(years=None):
    """Create a minimal valid workbook in memory."""
    if years is None:
        years = ["FY2022", "FY2023"]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_rows = {
        "P&L": ["Revenue", "COGS", "Operating Income", "Depreciation & Amortization", "Net Income"],
        "Balance Sheet": ["Current Assets", "Current Liabilities", "Total Assets", "Total Debt", "Total Shareholders Equity"],
        "Cash Flow": ["Operating Cash Flow", "Capital Expenditures"],
    }
    for sheet_name, labels in sheet_rows.items():
        ws = wb.create_sheet(sheet_name)
        ws.append([""] + years)          # header row: blank + year labels
        for label in labels:
            ws.append([label] + [100.0 + i * 10 for i in range(len(years))])
    return wb


def test_three_sheet_parse(tmp_path):
    """INPT-01: Parser reads P&L, Balance Sheet, Cash Flow sheets."""
    wb = _build_minimal_workbook()
    path = tmp_path / "test.xlsx"
    wb.save(path)
    result = parse_workbook(path)
    assert isinstance(result, FinancialData)
    assert "Revenue" in result.pl.rows
    assert result.pl.years == ["FY2022", "FY2023"]


def test_accepts_path(tmp_path):
    """INPT-03: parse_workbook accepts pathlib.Path input."""
    wb = _build_minimal_workbook()
    path = tmp_path / "test.xlsx"
    wb.save(path)
    result = parse_workbook(path)  # pathlib.Path
    assert isinstance(result, FinancialData)


def test_accepts_binary_io(tmp_path):
    """INPT-03: parse_workbook accepts BinaryIO stream (Streamlit use case)."""
    wb = _build_minimal_workbook()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = parse_workbook(buf)  # BinaryIO
    assert isinstance(result, FinancialData)
    assert result.company_name == "Unknown"


def test_missing_sheet_error(tmp_path):
    """INPT-01: Parser raises ValidationError when a required sheet is absent."""
    wb = openpyxl.Workbook()
    wb.active.title = "P&L"  # only one sheet, missing Balance Sheet and Cash Flow
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    with pytest.raises(ValidationError) as exc_info:
        parse_workbook(path)
    assert any("Balance Sheet" in e for e in exc_info.value.errors)


def test_demo_file_loads():
    """INPT-04: demo/apple_financials.xlsx parses and computes all metrics without error."""
    from autopitch.metrics import compute_metrics

    demo_path = Path("demo/apple_financials.xlsx")
    assert demo_path.exists(), "demo/apple_financials.xlsx not found — run create_demo.py"

    data = parse_workbook(demo_path)
    assert data.company_name == "apple_financials"
    assert data.pl.years == ["FY2020", "FY2021", "FY2022", "FY2023", "FY2024"]
    assert "Revenue" in data.pl.rows

    metrics = compute_metrics(data)
    # Revenue growth FY2022 ≈ 7.79%
    assert abs(metrics.revenue_growth["FY2022"] - 7.79) < 0.5
    # All metrics non-null for FY2022 (full data year, positive equity)
    assert metrics.gross_margin.get("FY2022") is not None
    assert metrics.net_margin.get("FY2022") is not None
    assert metrics.free_cash_flow.get("FY2022") is not None
    assert metrics.debt_to_equity.get("FY2022") is not None
    # FY2023 and FY2024 have negative equity — D/E and ROE should be None
    assert metrics.debt_to_equity.get("FY2023") is None
    assert metrics.debt_to_equity.get("FY2024") is None
